from __future__ import annotations

from datetime import date, datetime, time
from pathlib import Path

from openpyxl import load_workbook

from .common import clean_text, extraction_failure


MAX_NONEMPTY_CELLS = 50_000


def _cell_text(value: object) -> str:
    """Convert an Excel cell value into stable text for semantic clustering."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, (date, time)):
        return value.isoformat()
    if isinstance(value, bool):
        return "true" if value else "false"
    return clean_text(str(value))


class XlsxExtractor:
    """Extract worksheet names and non-empty cells from modern XLSX files."""

    name = "openpyxl"
    version = "1"

    def extract(self, path: Path, max_pages: int):
        """Extract workbook text.

        For XLSX files, MAX_PAGES is interpreted as the maximum number of
        worksheets to process. A value of 0 processes all worksheets.
        Formulas are retained as formula text because cached calculated values
        are not guaranteed to exist in every workbook.
        """
        workbook = None
        try:
            workbook = load_workbook(
                filename=path,
                read_only=True,
                data_only=False,
                keep_links=False,
            )

            worksheets = list(workbook.worksheets)
            selected = worksheets if max_pages <= 0 else worksheets[:max_pages]

            parts: list[str] = []
            nonempty_cells = 0
            truncated = False
            sheets_with_text = 0

            for worksheet in selected:
                sheet_parts: list[str] = []

                for row in worksheet.iter_rows():
                    row_values: list[str] = []
                    for cell in row:
                        value = _cell_text(cell.value)
                        if not value:
                            continue

                        # Coordinates preserve basic spreadsheet structure and
                        # help distinguish labels, headers, and nearby values.
                        row_values.append(f"{cell.coordinate}: {value}")
                        nonempty_cells += 1

                        if nonempty_cells >= MAX_NONEMPTY_CELLS:
                            truncated = True
                            break

                    if row_values:
                        sheet_parts.append(" | ".join(row_values))

                    if truncated:
                        break

                if sheet_parts:
                    sheets_with_text += 1
                    parts.append(
                        f"Worksheet {worksheet.title}: " + " ".join(sheet_parts)
                    )

                if truncated:
                    break

            text = clean_text("\n".join(parts))
            metadata = {
                "worksheet_count": len(worksheets),
                "worksheets_extracted": len(selected),
                "worksheets_with_text": sheets_with_text,
                "nonempty_cells_extracted": nonempty_cells,
                "truncated": truncated,
                "max_nonempty_cells": MAX_NONEMPTY_CELLS,
            }

            if not text:
                return (
                    "",
                    metadata,
                    extraction_failure(
                        "NO_EXTRACTABLE_TEXT",
                        "The XLSX workbook contains no extractable cell values.",
                    ),
                )

            return text, metadata, None

        except Exception as exc:
            return (
                "",
                {},
                extraction_failure("XLSX_EXTRACTION_FAILED", str(exc)),
            )
        finally:
            if workbook is not None:
                workbook.close()
