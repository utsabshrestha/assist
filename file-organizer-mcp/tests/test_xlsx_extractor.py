from openpyxl import Workbook

from file_organizer_mcp.extractors import build_extractor_registry


def test_registry_contains_xlsx():
    assert ".xlsx" in build_extractor_registry()


def test_xlsx_extracts_sheet_names_values_and_formulas(tmp_path):
    path = tmp_path / "budget.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Annual Budget"
    sheet["A1"] = "Department"
    sheet["B1"] = "Amount"
    sheet["A2"] = "Accessibility"
    sheet["B2"] = 12500
    sheet["C2"] = "=B2*1.05"
    workbook.save(path)
    workbook.close()

    text, metadata, error = build_extractor_registry()[".xlsx"].extract(path, 0)

    assert error is None
    assert "Worksheet Annual Budget" in text
    assert "A1: Department" in text
    assert "A2: Accessibility" in text
    assert "B2: 12500" in text
    assert "C2: =B2*1.05" in text
    assert metadata["worksheet_count"] == 1
    assert metadata["nonempty_cells_extracted"] == 5


def test_xlsx_max_pages_limits_worksheets(tmp_path):
    path = tmp_path / "multi-sheet.xlsx"
    workbook = Workbook()
    workbook.active.title = "First"
    workbook.active["A1"] = "Included"
    second = workbook.create_sheet("Second")
    second["A1"] = "Excluded"
    workbook.save(path)
    workbook.close()

    text, metadata, error = build_extractor_registry()[".xlsx"].extract(path, 1)

    assert error is None
    assert "Included" in text
    assert "Excluded" not in text
    assert metadata["worksheet_count"] == 2
    assert metadata["worksheets_extracted"] == 1
